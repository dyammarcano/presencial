import logging
import shutil
from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from tkinter import Tk, Toplevel, Label, Button, simpledialog, Frame
from tkinter.ttk import Radiobutton
from typing import Tuple, List, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='presence_tracker.log'
)


# Configuration for UI appearance
@dataclass
class UIConfig:
    FONT_FAMILY: str = "Courier"
    FONT_SIZE: int = 12
    BUTTON_WIDTH: int = 10
    DIALOG_PADDING: int = 20
    WRAP_LENGTH: int = 350
    WINDOW_WIDTH: int = 300
    WINDOW_HEIGHT: int = 150


# Configuration for app data and thresholds
@dataclass
class AppConfig:
    FOLDER_NAME: str = "Precencial"
    EXCEL_FILENAME: str = "registros.xlsx"
    CONFIG_FILENAME: str = "config.txt"
    EXCEL_HEADERS: Tuple[str, ...] = ("data", "hora", "resposta", "observacao", "area")
    DEFAULT_GOAL: int = 8
    MAX_GOAL: int = 31
    EXTRA_LABEL: str = "extra"


# Possible responses
class ResponseType(Enum):
    YES = "S"
    NO = "N"


# Work areas
class Area(Enum):
    AG = "AG"
    CT = "CT"
    CEIC = "CEIC"
    OTHER = "OUTRO"


# Base class for all dialogs
class BaseDialog:
    def __init__(self):
        self.window = None
        self.frame = None
        self.ui_config = UIConfig()

    def create_window(self, title: str) -> None:
        self.window = Toplevel()
        self.window.title(title)
        self.window.resizable(False, False)
        self.window.attributes("-topmost", True)
        self._center_window()
        self.frame = Frame(self.window)
        self.frame.pack(pady=10)

    def _center_window(self) -> None:
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width - self.ui_config.WINDOW_WIDTH) // 2
        y = (screen_height - self.ui_config.WINDOW_HEIGHT) // 2
        self.window.geometry(f"+{x}+{y}")

    def create_label(self, text: str) -> Label:
        return Label(
            self.window,
            text=text,
            padx=self.ui_config.DIALOG_PADDING,
            pady=self.ui_config.DIALOG_PADDING,
            wraplength=self.ui_config.WRAP_LENGTH,
            font=(self.ui_config.FONT_FAMILY, self.ui_config.FONT_SIZE)
        )

    def create_button(self, text: str, command: Callable[[], None]) -> Button:
        return Button(
            self.window,
            text=text,
            command=command,
            width=self.ui_config.BUTTON_WIDTH,
            font=(self.ui_config.FONT_FAMILY, self.ui_config.FONT_SIZE)
        )

    def destroy(self) -> None:
        if self.window:
            self.window.destroy()


# Message dialog built on top of BaseDialog
class MessageDialog(BaseDialog):
    def show(self, title: str, message: str, kind: str = "info") -> None:
        try:
            self.create_window(title)
            self.create_label(message).pack()
            btn_text = "OK" if kind != "error" else "Fechar"
            self.create_button(btn_text, self.window.destroy).pack(pady=10)
            self.window.grab_set()
            self.window.wait_window()
        except Exception as e:
            logging.error(f"Error showing message dialog: {str(e)}")
            raise


# Presence manager handles file I/O and configuration
class PresenceManager:
    def __init__(self, config: AppConfig):
        self.config = config
        self.message_dialog = MessageDialog()
        self.data_folder = self._initialize_data_folder()
        self.excel_path = self.data_folder / config.EXCEL_FILENAME
        self.config_path = self.data_folder / config.CONFIG_FILENAME
        self.monthly_goal = self._load_or_setup_config()
        self._copy_self_to_prevalence()
        self._ensure_excel_exists()

    def _copy_self_to_prevalence(self) -> None:
        try:
            current_file = Path(__file__).resolve()
            destination = self.data_folder / current_file.name
            if not destination.exists():
                shutil.copy(current_file, destination)
                logging.info(f"Application copied to {destination}")
            else:
                logging.info(f"Application already exists at {destination}")
        except Exception as e:
            logging.error(f"Failed to copy application: {e}")

    def _initialize_data_folder(self) -> Path:
        home = Path.home()
        folder_path = home / self.config.FOLDER_NAME
        folder_path.mkdir(exist_ok=True)
        return folder_path

    def _ensure_excel_exists(self) -> None:
        if not self.excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(self.config.EXCEL_HEADERS)
            wb.save(self.excel_path)

    def save_presence(self, is_present: bool, observation: str = "", area: str = "") -> None:
        now = datetime.now()
        response = ResponseType.YES.value if is_present else ResponseType.NO.value
        wb = load_workbook(self.excel_path)
        ws = wb.active
        ws.append([
            str(now.date()),
            now.strftime("%H:%M:%S"),
            response,
            observation,
            area
        ])
        wb.save(self.excel_path)

    def count_monthly_presence(self) -> Tuple[int, List[str]]:
        current = datetime.now()
        total = 0
        records = []
        wb = load_workbook(self.excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                date = datetime.strptime(row[0], "%Y-%m-%d")
                if date.year == current.year and date.month == current.month and row[2] == ResponseType.YES.value:
                    total += 1
                    area = row[4] or 'N/A'
                    records.append(f"* {row[0]} Presencial no {area.ljust(6)}")
            except Exception as e:
                logging.warning(f"Error processing row {row}: {e}")
        return total, records

    def _load_or_setup_config(self) -> int:
        try:
            if self.config_path.exists():
                with open(self.config_path, "r", encoding='utf-8') as file:
                    value = int(file.read().strip())
                    if 1 <= value <= self.config.MAX_GOAL:
                        return value
            goal = simpledialog.askinteger(
                "Configuração Inicial",
                "Quantos dias presenciais por mês você deseja atingir?",
                minvalue=1,
                maxvalue=self.config.MAX_GOAL
            )
            if not goal:
                goal = self.config.DEFAULT_GOAL
                self.message_dialog.show(
                    "Erro",
                    f"Valor inválido. Usando valor padrão de {self.config.DEFAULT_GOAL}.",
                    "error"
                )
            with open(self.config_path, "w", encoding='utf-8') as file:
                file.write(str(goal))
            return goal
        except Exception as e:
            logging.error(f"Error loading/setting up config: {str(e)}")
            return self.config.DEFAULT_GOAL


# UI class handling user interaction and flow
class PresenceUI:
    def __init__(self, manager: PresenceManager):
        self.manager = manager
        self.message_dialog = MessageDialog()

    def _show_summary(self, total: int, records: List[str]) -> None:
        # Display a summary of presence records for the current month
        summary = f"Você já marcou presença {total} vez(es) neste mês."
        if records:
            summary += "\n\n" + "\n".join(records)
        self.message_dialog.show("Resumo do mês", summary)

    def ask_is_present(self) -> bool:
        # Ask user if they are present today
        dialog = BaseDialog()
        result = [False]

        def on_yes():
            result[0] = True
            dialog.destroy()

        def on_no():
            result[0] = False
            dialog.destroy()

        dialog.create_window("Precencial")
        dialog.create_label("Você está presencial hoje?").pack()

        frame = Frame(dialog.window)
        frame.pack(pady=10)
        dialog.create_button("Sim", on_yes).pack(in_=frame, side="left", padx=5)
        dialog.create_button("Não", on_no).pack(in_=frame, side="right", padx=5)

        dialog.window.grab_set()
        dialog.window.wait_window()
        return result[0]

    def _ask_extra_confirmation(self) -> bool:
        # Ask a user if they want to record extra presence after reaching the goal
        return self.ask_yes_no(
            "Meta Atingida",
            f"Você já registrou {self.manager.monthly_goal} vezes este mês.\nDeseja registrar mais uma por conta própria?"
        )

    def ask_yes_no(self, title: str, question: str) -> bool:
        # General purpose Yes/No dialog
        dialog = BaseDialog()
        result = [False]

        def on_yes():
            result[0] = True
            dialog.destroy()

        def on_no():
            result[0] = False
            dialog.destroy()

        dialog.create_window(title)
        dialog.create_label(question).pack()

        frame = Frame(dialog.window)
        frame.pack(pady=10)
        dialog.create_button("Sim", on_yes).pack(in_=frame, side="left", padx=5)
        dialog.create_button("Não", on_no).pack(in_=frame, side="right", padx=5)

        dialog.window.grab_set()
        dialog.window.wait_window()
        return result[0]

    def _ask_area_selection_and_save(self, observation: str) -> None:
        # Area selection dialog with save operation
        from tkinter import StringVar

        dialog = BaseDialog()
        var = StringVar()

        def on_select():
            selected = var.get()
            if selected:
                dialog.destroy()
                self.manager.save_presence(True, observation, selected)
                MessageDialog().show("Registrado", "Sua resposta foi salva com sucesso.")
            else:
                error_label.config(text="Selecione uma área.", fg="red")

        dialog.create_window("Área")
        dialog.create_label("Em qual área você está hoje?").pack()

        for area in Area:
            Radiobutton(dialog.window, text=area.value, variable=var, value=area.value).pack(anchor="w", padx=20)

        error_label = Label(dialog.window, text="", font=("Courier", 12))
        error_label.pack()

        Button(dialog.window, text="Confirmar", command=on_select, width=15, font=("Courier", 12)).pack(pady=10)

        dialog.window.grab_set()
        dialog.window.wait_window()

    def ask_presence(self) -> None:
        # Overall flow for presence interaction
        total, records = self.manager.count_monthly_presence()
        self._show_summary(total, records)

        if self.ask_is_present():
            self._handle_present_response(total)
        else:
            self._handle_absent_response()

    def _handle_present_response(self, total: int) -> None:
        # Handles logic for user marked as present
        if total >= self.manager.monthly_goal:
            if not self._ask_extra_confirmation():
                self.message_dialog.show("Encerrado", "Nenhum registro foi adicionado.")
                return
            observation = self.manager.config.EXTRA_LABEL
        else:
            observation = ""
        self._ask_area_selection_and_save(observation)

    def _handle_absent_response(self) -> None:
        # Handles logic for user marked as absent
        self.manager.save_presence(False)
        self.message_dialog.show("Informativo", "Tudo bem. Hoje não será contado como presencial.")


# Application entry point
def main():
    root = None
    try:
        root = Tk()
        root.withdraw()

        config = AppConfig()
        manager = PresenceManager(config)
        ui = PresenceUI(manager)
        ui.ask_presence()
    except Exception as e:
        logging.error(f"Application error: {str(e)}")
        MessageDialog().show("Erro", f"Ocorreu um erro: {str(e)}", "error")
    finally:
        if root and root.winfo_exists():
            root.destroy()


if __name__ == "__main__":
    main()
